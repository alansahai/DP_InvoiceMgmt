import streamlit as st
import pandas as pd
import time
import json
from datetime import datetime, timedelta
import processor
import math
from io import BytesIO
from urllib.parse import urlparse
from urllib.request import urlopen, Request
from openpyxl.utils import get_column_letter
from mail_ingestion import ingest_invoices_from_email, is_mail_ingestion_configured
from database import (
    upload_file, 
    save_invoice_record, 
    fetch_all_invoices, 
    fetch_all_invoice_edits,
    fetch_all_invoice_audits,
    fetch_all_vendors,
    is_duplicate, 
    log_edit,
    get_vendor_average,
    fetch_invoice_edits,
    compute_document_hash,
    is_duplicate_hash
)

st.set_page_config(page_title="AI Invoice Auditor", layout="wide")

# --- CONFIG ---
CURRENT_AI_VERSION = "gemini-flash-lite-latest"

# --- 🔐 RBAC: LOGIN SIMULATION ---
with st.sidebar:
    st.header("👤 User Identity")
    user_role = st.selectbox("Login As:", ["AP_CLERK", "FINANCE_MANAGER", "AUDITOR"], index=0)
    st.caption(f"Permissions: {user_role}")
    st.markdown("---")
    
    # API Key Status Display
    st.header("🔑 API Status")
    try:
        st.caption(f"Total Keys: {len(processor.API_KEYS)}")
        for i in range(len(processor.API_KEYS)):
            if i in processor.failed_keys:
                st.error(f"Key #{i+1}: ❌ Quota Exceeded")
            elif i == processor.current_key_index:
                st.success(f"Key #{i+1}: ✅ Active")
            else:
                st.info(f"Key #{i+1}: 💤 Standby")
        st.caption(f"Next: Key #{processor.current_key_index + 1}")
    except:
        st.warning("API Status Unavailable")
    st.markdown("---")

def can_edit(): return user_role in ["AP_CLERK", "FINANCE_MANAGER"]
def can_upload(): return user_role == "AP_CLERK"
def can_approve(): return user_role == "FINANCE_MANAGER"

st.title("🛡️ AI-Powered Invoice Auditor")

# --- 🎨 VISUAL BADGE MAPPING ---
def get_stage_badge(stage):
    badges = {
        "UPLOADED": "🔵 Uploaded",
        "REVIEWED": "🟠 Reviewed",
        "APPROVED": "🟢 Approved",
        "REJECTED": "🔴 Rejected",
        "AUDITED": "🟣 Audited"
    }
    return badges.get(stage, f"⚪ {stage}")

def get_risk_badge(risk_level):
    badges = {
        "LOW": "✅ Low Risk",
        "MEDIUM": "⚠️ Medium Risk",
        "HIGH": "🚨 High Risk"
    }
    return badges.get(risk_level, risk_level)

def get_processing_badge(processing_status):
    badges = {
        "MANUAL_UPLOAD": "🖐️ Manual Upload",
        "INGESTED_EMAIL": "📩 IMAP Ingested",
        "COMPLETED": "✅ Processed"
    }
    return badges.get(processing_status or "COMPLETED", processing_status or "COMPLETED")

def get_invoice_source(record):
    created_by = str(record.get("created_by") or "").upper()
    processing_status = str(record.get("processing_status") or "").upper()
    if created_by == "MAIL_BOT" or "INGEST" in processing_status:
        return "INGESTED_EMAIL"
    return "MANUAL_UPLOAD"

def hydrate_invoice_session_data(row):
    invoice_data = row.get('ai_raw_data') or {}
    invoice_data['id'] = row.get('id')
    invoice_data['created_by'] = row.get('created_by')
    invoice_data['approval_stage'] = row.get('approval_stage')
    invoice_data['reviewed_by'] = row.get('reviewed_by')
    invoice_data['approved_by'] = row.get('approved_by')
    invoice_data['approval_timestamp'] = row.get('approval_timestamp')
    invoice_data['processing_status'] = row.get('processing_status')
    invoice_data['document_hash'] = row.get('document_hash')
    invoice_data['risk_level'] = row.get('risk_level')
    invoice_data['risk_score'] = row.get('risk_score')
    invoice_data['flag_reason'] = row.get('flag_reason')
    invoice_data['confidence_score'] = row.get('confidence_score')
    invoice_data['ai_version'] = row.get('ai_version') or CURRENT_AI_VERSION
    return invoice_data

# --- 🧠 Logic Engine ---
def validate_math(invoice_total, line_items_df):
    if line_items_df.empty: return False, 0.0, invoice_total
    calculated_sum = line_items_df['total_price'].sum()
    diff = round(abs(invoice_total - calculated_sum), 2)
    return diff == 0, calculated_sum, diff

# --- 🧹 JSON SANITIZATION HELPER ---
def sanitize_json(obj):
    """Recursively replace NaN/Inf with None for JSON compatibility."""
    if isinstance(obj, dict):
        return {k: sanitize_json(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [sanitize_json(v) for v in obj]
    elif isinstance(obj, float):
        if math.isnan(obj) or math.isinf(obj):
            return None
    return obj

def _is_pdf_url(url):
    if not url:
        return False
    parsed = urlparse(str(url))
    path = (parsed.path or "").lower()
    return path.endswith(".pdf")

def render_document_preview(file_url):
    if not file_url:
        st.warning("No document URL available for this invoice.")
        return

    st.markdown(f"[Open File]({file_url})")

    try:
        request = Request(file_url, headers={"User-Agent": "Mozilla/5.0"})
        with urlopen(request, timeout=15) as response:
            file_bytes = response.read()
            content_type = str(response.headers.get("Content-Type") or "").lower()

        if "pdf" in content_type or _is_pdf_url(file_url):
            st.info("PDF preview is opened via link above.")
            return

        if "image" in content_type:
            st.image(file_bytes, use_container_width=True)
            return

        try:
            st.image(file_bytes, use_container_width=True)
            return
        except Exception:
            st.warning("Preview unavailable for this file type.")
            return

    except Exception:
        try:
            st.image(file_url, use_container_width=True)
        except Exception:
            st.warning("Preview unavailable for this invoice file.")


def _serialize_export_value(value):
    if isinstance(value, (dict, list)):
        try:
            return json.dumps(value, ensure_ascii=False)
        except Exception:
            return str(value)
    return value


def _prepare_export_df(records):
    if not records:
        return pd.DataFrame()

    serialized_records = []
    for record in records:
        row = {}
        for key, value in (record or {}).items():
            row[key] = _serialize_export_value(value)
        serialized_records.append(row)

    return pd.DataFrame(serialized_records)


def _autofit_worksheet(worksheet, dataframe):
    if dataframe.empty:
        return

    for idx, col in enumerate(dataframe.columns):
        try:
            max_value_len = dataframe[col].astype(str).str.len().max()
            if pd.isna(max_value_len):
                max_value_len = 0
            max_length = max(max_value_len, len(str(col))) + 2
            col_letter = get_column_letter(idx + 1)
            worksheet.column_dimensions[col_letter].width = min(max_length, 60)
        except Exception:
            continue


def export_full_transparency_workbook(invoices, edits, audits, vendors):
    output = BytesIO()

    invoices_df = _prepare_export_df(invoices)
    edits_df = _prepare_export_df(edits)
    audits_df = _prepare_export_df(audits)
    vendors_df = _prepare_export_df(vendors)

    workflow_df = pd.DataFrame()
    if not invoices_df.empty:
        source_series = invoices_df.apply(
            lambda row: get_invoice_source(row.to_dict()),
            axis=1
        )
        stage_series = invoices_df.get('approval_stage', pd.Series(dtype='object'))
        workflow_df = pd.DataFrame({
            "invoice_id": invoices_df.get('id', pd.Series(dtype='object')),
            "vendor_name": invoices_df.get('vendor_name', pd.Series(dtype='object')),
            "source": source_series,
            "approval_stage": stage_series,
            "status": invoices_df.get('status', pd.Series(dtype='object')),
            "processing_status": invoices_df.get('processing_status', pd.Series(dtype='object')),
            "risk_level": invoices_df.get('risk_level', pd.Series(dtype='object')),
            "risk_score": invoices_df.get('risk_score', pd.Series(dtype='object')),
            "created_by": invoices_df.get('created_by', pd.Series(dtype='object')),
            "reviewed_by": invoices_df.get('reviewed_by', pd.Series(dtype='object')),
            "approved_by": invoices_df.get('approved_by', pd.Series(dtype='object')),
            "audited": invoices_df.get('audited', pd.Series(dtype='object')),
            "created_at": invoices_df.get('created_at', pd.Series(dtype='object')),
            "approval_timestamp": invoices_df.get('approval_timestamp', pd.Series(dtype='object')),
        })

    line_items_rows = []
    for invoice in invoices or []:
        line_items = []
        if isinstance(invoice.get("ai_raw_data"), dict):
            line_items = invoice.get("ai_raw_data", {}).get("line_items", []) or []

        if not line_items and isinstance(invoice.get("ai_structured_output"), dict):
            line_items = invoice.get("ai_structured_output", {}).get("line_items", []) or []

        for idx, item in enumerate(line_items, start=1):
            line_items_rows.append({
                "invoice_id": invoice.get("id"),
                "vendor_name": invoice.get("vendor_name"),
                "line_number": idx,
                "description": _serialize_export_value(item.get("description") if isinstance(item, dict) else item),
                "quantity": _serialize_export_value(item.get("quantity") if isinstance(item, dict) else None),
                "unit_price": _serialize_export_value(item.get("unit_price") if isinstance(item, dict) else None),
                "total_price": _serialize_export_value(item.get("total_price") if isinstance(item, dict) else None),
            })
    line_items_df = pd.DataFrame(line_items_rows)

    ai_flat_df = pd.DataFrame()
    try:
        ai_records = []
        for invoice in invoices or []:
            ai_record = invoice.get("ai_raw_data") if isinstance(invoice.get("ai_raw_data"), dict) else {}
            ai_record = ai_record.copy()
            ai_record["invoice_id"] = invoice.get("id")
            ai_records.append(ai_record)

        if ai_records:
            ai_flat_df = pd.json_normalize(ai_records, sep='.')
            for col in ai_flat_df.columns:
                ai_flat_df[col] = ai_flat_df[col].apply(_serialize_export_value)
    except Exception:
        ai_flat_df = pd.DataFrame()

    stage_counts = {}
    source_counts = {}
    if not workflow_df.empty:
        if 'approval_stage' in workflow_df.columns:
            stage_counts = workflow_df['approval_stage'].value_counts(dropna=False).to_dict()
        if 'source' in workflow_df.columns:
            source_counts = workflow_df['source'].value_counts(dropna=False).to_dict()

    summary_rows = [
        {"metric": "generated_at", "value": datetime.now().isoformat()},
        {"metric": "total_invoices", "value": len(invoices or [])},
        {"metric": "total_invoice_edits", "value": len(edits or [])},
        {"metric": "total_invoice_audits", "value": len(audits or [])},
        {"metric": "total_vendors", "value": len(vendors or [])},
    ]
    for stage, count in stage_counts.items():
        summary_rows.append({"metric": f"stage_{stage}", "value": count})
    for source, count in source_counts.items():
        summary_rows.append({"metric": f"source_{source}", "value": count})
    summary_df = pd.DataFrame(summary_rows)

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        invoices_df.to_excel(writer, sheet_name='Invoices_Full', index=False)
        workflow_df.to_excel(writer, sheet_name='Workflow_View', index=False)
        line_items_df.to_excel(writer, sheet_name='Line_Items', index=False)
        edits_df.to_excel(writer, sheet_name='Invoice_Edits', index=False)
        audits_df.to_excel(writer, sheet_name='Invoice_Audits', index=False)
        vendors_df.to_excel(writer, sheet_name='Vendors', index=False)

        if not ai_flat_df.empty:
            ai_flat_df.to_excel(writer, sheet_name='AI_Flattened', index=False)

        for sheet_name, worksheet in writer.sheets.items():
            if sheet_name == 'Summary':
                _autofit_worksheet(worksheet, summary_df)
            elif sheet_name == 'Invoices_Full':
                _autofit_worksheet(worksheet, invoices_df)
            elif sheet_name == 'Workflow_View':
                _autofit_worksheet(worksheet, workflow_df)
            elif sheet_name == 'Line_Items':
                _autofit_worksheet(worksheet, line_items_df)
            elif sheet_name == 'Invoice_Edits':
                _autofit_worksheet(worksheet, edits_df)
            elif sheet_name == 'Invoice_Audits':
                _autofit_worksheet(worksheet, audits_df)
            elif sheet_name == 'Vendors':
                _autofit_worksheet(worksheet, vendors_df)
            elif sheet_name == 'AI_Flattened':
                _autofit_worksheet(worksheet, ai_flat_df)

    output.seek(0)
    return output.getvalue()

# --- 📊 EXCEL EXPORT HELPER ---
def export_to_excel(dataframe, filename="export"):
    """Convert DataFrame to Excel file and return as bytes."""
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        dataframe.to_excel(writer, sheet_name='Invoices', index=False)
        
        # Auto-format columns
        workbook = writer.book
        worksheet = writer.sheets['Invoices']
        for idx, col in enumerate(dataframe.columns):
            max_value_len = dataframe[col].astype(str).str.len().max()
            if pd.isna(max_value_len):
                max_value_len = 0
            max_length = max(max_value_len, len(str(col))) + 2
            col_letter = get_column_letter(idx + 1)
            worksheet.column_dimensions[col_letter].width = min(max_length, 50)
    
    output.seek(0)
    return output.getvalue()

# --- 📊 INVOICE EXPORT WITH LINE ITEMS ---
def export_invoice_with_items(invoice_summary, line_items_df):
    """Export invoice summary and line items to Excel with multiple sheets."""
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Sheet 1: Invoice Summary
        summary_df = pd.DataFrame([invoice_summary])
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        # Sheet 2: Line Items (if available)
        if not line_items_df.empty:
            line_items_df.to_excel(writer, sheet_name='Line Items', index=False)
        
        # Auto-format columns for all sheets
        workbook = writer.book
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            # Get the dataframe for this sheet
            if sheet_name == 'Summary':
                df = summary_df
            else:
                df = line_items_df
            
            # Format column widths
            for idx, col in enumerate(df.columns):
                try:
                    max_value_len = df[col].astype(str).str.len().max()
                    if pd.isna(max_value_len):
                        max_value_len = 0
                    max_length = max(max_value_len, len(str(col))) + 2
                    col_letter = get_column_letter(idx + 1)
                    worksheet.column_dimensions[col_letter].width = min(max_length, 50)
                except:
                    pass
    
    output.seek(0)
    return output.getvalue()

# --- SAVE HELPER ---
def save_and_log(vendor, date, total, cur, df, status, original_data, reason, role, risk_score, risk_level, stage):
    approval_ts = None
    reviewed_by = None
    approved_by = None
    
    if stage == "REVIEWED":
        reviewed_by = role
    elif stage == "APPROVED":
        reviewed_by = original_data.get("reviewed_by") or role
        approved_by = role
        approval_ts = datetime.now().isoformat()
    elif stage == "REJECTED":
        reviewed_by = role
        approved_by = None

    # Preserve edited line_items
    updated_raw_data = original_data.copy()
    updated_raw_data["line_items"] = df.to_dict("records")
    
    final_data = {
        "vendor_name": vendor, "invoice_date": date, "total_amount": total,
        "currency": cur, "line_items": df.to_dict("records"),
        "validation_status": status,
        "processing_status": original_data.get("processing_status", "MANUAL_UPLOAD"),
        "confidence_score": original_data.get("confidence_score", 0.0),
        "flag_reason": reason, 
        "document_hash": original_data.get("document_hash"),
        "ai_raw_data": updated_raw_data,
        "ai_structured_output": original_data.get("ai_raw_structured"),
        "ai_explanations": original_data.get("explanations"),
        "risk_score": risk_score, "risk_level": risk_level,
        "approval_stage": stage, "reviewed_by": reviewed_by,
        "approved_by": approved_by, "approval_timestamp": approval_ts,
        "ai_version": original_data.get("ai_version", CURRENT_AI_VERSION),
        "reprocessed_at": original_data.get("reprocessed_at"),
        "audited": True if stage == "AUDITED" else original_data.get("audited", False)
    }
    
    if "created_by" in original_data:
        final_data["created_by"] = original_data["created_by"]
    
    # ✅ FIX: Extract ID correctly to ensure UPDATE instead of INSERT
    invoice_id = original_data.get("id")
    
    clean_final_data = sanitize_json(final_data)
    saved = save_invoice_record(clean_final_data, st.session_state['url'], role, invoice_id=invoice_id)
    
    if saved:
        invoice_id = saved['id']
        if vendor != original_data.get("vendor_name"):
            log_edit(invoice_id, "Vendor", original_data.get("vendor_name"), vendor)
        if str(total) != str(original_data.get("total_amount")):
            log_edit(invoice_id, "Total", original_data.get("total_amount"), total)
        if date != original_data.get("invoice_date"):
            log_edit(invoice_id, "Date", original_data.get("invoice_date"), date)
            
        original_items = pd.DataFrame(original_data.get("line_items", []))
        if not original_items.empty and not df.equals(original_items):
            log_edit(invoice_id, "Line Items", "Original AI Table", "User Modified Table")
            
        st.toast(f"Invoice moved to {stage} stage!", icon="✅")
        time.sleep(1)
        del st.session_state['data']
        st.rerun()
    else:
        st.error("Action blocked by workflow policy. Please check role permissions and stage order.")

# --- HELPER: FETCH VENDOR STATS ---
def get_vendor_stats(vendor_name, all_invoices):
    if not all_invoices or not vendor_name:
        return None
    
    v_df = pd.DataFrame(all_invoices)
    v_data = v_df[v_df['vendor_name'] == vendor_name]
    
    if v_data.empty:
        return None
        
    v_data['invoice_date'] = pd.to_datetime(v_data['invoice_date'], errors='coerce')
        
    stats = {
        "count": len(v_data),
        "avg_amount": v_data['total_amount'].mean(),
        "last_invoice": v_data['invoice_date'].max(),
        "flagged_count": len(v_data[v_data['risk_level'] == 'HIGH']) if 'risk_level' in v_data.columns else 0
    }
    return stats

# --- Sidebar: Ingestion ---
with st.sidebar:
    st.header("1. Upload Invoice")
    if can_upload():
        uploaded_file = st.file_uploader("Upload File", type=["pdf", "png", "jpg", "jpeg"])
        if uploaded_file and st.button("Analyze Invoice"):
            st.session_state.pop('data', None)
            st.session_state.pop('file_bytes', None)
            st.session_state.pop('url', None)
            
            with st.spinner(f"🔍 AI ({CURRENT_AI_VERSION}) is processing..."):
                file_bytes = uploaded_file.getvalue()
                document_hash = compute_document_hash(file_bytes)
                if is_duplicate_hash(document_hash):
                    st.warning("Potential duplicate detected (same document hash found).")
                public_url = upload_file(file_bytes, uploaded_file.name, uploaded_file.type)
                if public_url:
                    data = processor.process_invoice(file_bytes, uploaded_file.type)
                    if not data:
                        st.error("AI processing failed. Please retry.")
                        st.stop()
                        
                    if data:
                        data['ai_version'] = CURRENT_AI_VERSION
                        data['processing_status'] = "MANUAL_UPLOAD"
                        data['document_hash'] = document_hash
                        st.session_state['data'] = data
                        st.session_state['url'] = public_url
                        st.session_state['file_bytes'] = file_bytes
                        st.session_state['mime_type'] = uploaded_file.type
                        st.rerun()
    else:
        st.info("ℹ️ Only AP Clerk can upload invoices")

    st.markdown("---")
    st.header("2. Mailbox Ingestion")
    if can_upload():
        configured, config_msg = is_mail_ingestion_configured()
        if configured:
            st.success("✅ IMAP Configured")
            max_messages = st.number_input("Emails to scan", min_value=1, max_value=100, value=20, step=1)
            if st.button("📩 Check Mailbox & Ingest"):
                with st.spinner("Checking mailbox and ingesting invoices..."):
                    st.session_state['last_ingestion_result'] = ingest_invoices_from_email(
                        max_messages=int(max_messages),
                        ai_version=CURRENT_AI_VERSION
                    )
        else:
            st.warning(f"IMAP unavailable: {config_msg}")
    else:
        st.info("ℹ️ Only AP Clerk can run mailbox ingestion")

    if st.session_state.get('last_ingestion_result'):
        ingest = st.session_state['last_ingestion_result']
        st.caption(
            f"Last run: {ingest.get('status')} | Ingested {ingest.get('ingested', 0)} | "
            f"Duplicates {ingest.get('duplicates', 0)} | Failed {ingest.get('failed', 0)}"
        )
        st.caption(
            f"Scanned {ingest.get('messages_scanned', 0)} mails | "
            f"Mails with attachments {ingest.get('messages_with_attachments', 0)} | "
            f"Attachments found {ingest.get('attachments_found', 0)}"
        )
        if ingest.get('skipped_by_type', 0) or ingest.get('skipped_by_size', 0):
            st.caption(
                f"Skipped by type: {ingest.get('skipped_by_type', 0)} | "
                f"Skipped by size: {ingest.get('skipped_by_size', 0)}"
            )
        errors = ingest.get('errors', [])[:3]
        if errors:
            st.warning("Ingestion errors:\n- " + "\n- ".join(errors))

# --- Pre-fetch invoices ---
all_invoices_data = fetch_all_invoices()

# --- 🔎 WORKFLOW TRANSPARENCY (ALL ROLES) ---
if all_invoices_data:
    st.markdown("---")
    st.markdown("### 🔎 Workflow Transparency")

    transparency_df = pd.DataFrame(all_invoices_data)
    transparency_df['source'] = transparency_df.apply(get_invoice_source, axis=1)

    total_invoices = len(transparency_df)
    active_invoices = len(transparency_df[transparency_df['approval_stage'].isin(["UPLOADED", "REVIEWED"])]) if 'approval_stage' in transparency_df.columns else 0
    approved_invoices = len(transparency_df[transparency_df['approval_stage'] == "APPROVED"]) if 'approval_stage' in transparency_df.columns else 0
    rejected_invoices = len(transparency_df[transparency_df['approval_stage'] == "REJECTED"]) if 'approval_stage' in transparency_df.columns else 0

    w1, w2, w3, w4 = st.columns(4)
    w1.metric("Total", total_invoices)
    w2.metric("Active", active_invoices)
    w3.metric("Approved", approved_invoices)
    w4.metric("Rejected", rejected_invoices)

    if 'approval_stage' in transparency_df.columns:
        source_stage = transparency_df.groupby(['source', 'approval_stage']).size().unstack(fill_value=0)
        st.caption("Source vs Approval Stage")
        st.dataframe(source_stage, use_container_width=True)

    st.markdown("### 📥 Full Transparency Export (All Users)")
    all_edits = fetch_all_invoice_edits()
    all_audits = fetch_all_invoice_audits()
    all_vendors = fetch_all_vendors()
    transparency_excel = export_full_transparency_workbook(
        all_invoices_data,
        all_edits,
        all_audits,
        all_vendors
    )
    st.download_button(
        label="📊 Download Full Transparency Workbook (EXCEL)",
        data=transparency_excel,
        file_name=f"invoice_transparency_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    st.caption("Includes invoices, workflow view, line items, invoice edits, invoice audits, vendors, and flattened AI extraction fields.")

# --- 📊 SYSTEM HEALTH SUMMARY ---
if all_invoices_data:
    st.markdown("---")
    st.markdown("### 🏥 System Health Summary")
    
    df_health = pd.DataFrame(all_invoices_data)
    
    try:
        api_keys_remaining = len(processor.API_KEYS) - len(processor.failed_keys)
    except:
        api_keys_remaining = 0
    
    avg_confidence = df_health['confidence_score'].mean() if 'confidence_score' in df_health.columns and not df_health.empty else 0.0
    
    high_risk_pending = len(df_health[
        (df_health['risk_level'] == 'HIGH') & 
        (df_health['approval_stage'].isin(["UPLOADED", "REVIEWED"]))
    ]) if 'risk_level' in df_health.columns and not df_health.empty else 0
    
    avg_approval_time = 0
    if 'approval_timestamp' in df_health.columns and 'created_at' in df_health.columns:
        approved_df = df_health[df_health['approval_stage'] == 'APPROVED'].copy()
        if not approved_df.empty:
            approved_df['approval_timestamp'] = pd.to_datetime(approved_df['approval_timestamp'], errors='coerce')
            approved_df['created_at'] = pd.to_datetime(approved_df['created_at'], errors='coerce')
            approved_df['approval_hours'] = (approved_df['approval_timestamp'] - approved_df['created_at']).dt.total_seconds() / 3600
            avg_approval_time = approved_df['approval_hours'].mean() if not approved_df.empty else 0
    
    h1, h2, h3, h4 = st.columns(4)
    try:
        h1.metric("🔑 API Keys Available", f"{api_keys_remaining}/{len(processor.API_KEYS)}", delta=f"{len(processor.failed_keys)} exhausted")
    except:
        h1.metric("🔑 API Keys Available", "N/A")
    
    h2.metric("📊 Avg AI Confidence", f"{avg_confidence:.1%}", delta="Quality Score")
    h3.metric("🚨 High-Risk Pending", f"{high_risk_pending}", delta="Require Attention")
    h4.metric("⏱ Avg Approval Time", f"{avg_approval_time:.1f} hrs", delta="Target: <48hrs")

# --- Main Dashboard ---
if 'data' in st.session_state:
    data = st.session_state['data']
    col1, col2 = st.columns([1, 1.5])
    
    with col1:
        st.subheader("Original Document")
        render_document_preview(st.session_state.get('url'))
            
        if can_edit() and 'file_bytes' in st.session_state:
            st.markdown("---")
            st.caption(f"Current Model: {data.get('ai_version', 'Unknown')}")
            if st.button("🔄 Reprocess with Latest AI"):
                with st.spinner(f"Re-running analysis with {CURRENT_AI_VERSION}..."):
                    new_data = processor.process_invoice(st.session_state['file_bytes'], st.session_state['mime_type'])
                    if new_data:
                        new_data['ai_version'] = CURRENT_AI_VERSION
                        new_data['reprocessed_at'] = datetime.now().isoformat()
                        new_data['processing_status'] = data.get('processing_status', 'MANUAL_UPLOAD')
                        new_data['document_hash'] = data.get('document_hash')
                        new_data.setdefault("explanations", {})
                        
                        # Preserve ID if reprocessing an existing invoice
                        if data.get('id'):
                            new_data['id'] = data.get('id')
                            
                        st.session_state['data'] = new_data
                        st.success("Invoice re-analyzed!")
                        time.sleep(1)
                        st.rerun()
    
    with col2:
        st.subheader("📝 Audit Results")
        status_col1, status_col2, status_col3 = st.columns(3)
        status_col1.caption(f"Stage: {get_stage_badge(data.get('approval_stage', 'UPLOADED'))}")
        status_col2.caption(f"Source: {get_processing_badge(data.get('processing_status'))}")
        status_col3.caption(f"Created By: {data.get('created_by', 'Unknown')}")
        
        # ✅ DOWNLOAD CURRENT INVOICE AS CSV/EXCEL
        down_col1, down_col2 = st.columns(2)
        
        # Prepare invoice summary data for download
        invoice_export = {
            "Vendor": data.get("vendor_name"),
            "Invoice Date": data.get("invoice_date"),
            "Total Amount ($)": data.get("total_amount"),
            "Currency": data.get("currency"),
            "Validation Status": data.get("validation_status"),
            "Risk Level": data.get("risk_level"),
            "Risk Score": data.get("risk_score"),
            "AI Confidence": data.get("confidence_score"),
            "Flag Reason": data.get("flag_reason"),
            "Approval Stage": data.get("approval_stage"),
            "Reviewed By": data.get("reviewed_by"),
            "Approved By": data.get("approved_by"),
            "AI Version": data.get("ai_version", CURRENT_AI_VERSION)
        }
        
        # Prepare line items for Excel export
        line_items_df = pd.DataFrame(data.get("line_items", []))
        
        with down_col1:
            csv_data = pd.DataFrame([invoice_export]).to_csv(index=False).encode('utf-8')
            st.download_button(
                label="📥 Download as CSV",
                data=csv_data,
                file_name=f"invoice_{data.get('vendor_name', 'unknown').replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.csv",
                mime="text/csv"
            )
        
        with down_col2:
            excel_data = export_invoice_with_items(invoice_export, line_items_df)
            st.download_button(
                label="📊 Download as EXCEL",
                data=excel_data,
                file_name=f"invoice_{data.get('vendor_name', 'unknown').replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
        st.divider()
        
        # --- 3️⃣ VENDOR INSIGHTS PANEL ---
        vendor_name = data.get("vendor_name")
        if vendor_name:
            v_stats = get_vendor_stats(vendor_name, all_invoices_data)
            if v_stats:
                with st.expander(f"🏢 Vendor Insights: {vendor_name}", expanded=False):
                    vc1, vc2, vc3, vc4 = st.columns(4)
                    vc1.metric("Total Invoices", v_stats['count'])
                    vc2.metric("Avg Amount", f"${v_stats['avg_amount']:,.2f}")
                    vc3.metric("Flagged History", v_stats['flagged_count'])
                    vc4.write(f"**Last Seen:**\n{v_stats['last_invoice']}")

        is_locked = data.get("approval_stage") == "APPROVED"
        
        with st.container(border=True):
            st.caption("Invoice Metadata")
            c1, c2 = st.columns(2)
            vendor = c1.text_input("Vendor", data.get("vendor_name"), disabled=not can_edit() or is_locked)
            date = c2.text_input("Date", data.get("invoice_date"), disabled=not can_edit() or is_locked)
            c3, c4 = st.columns(2)
            extracted_total = c3.number_input("Invoice Total ($)", value=float(data.get("total_amount", 0.0)), disabled=not can_edit() or is_locked)
            currency = c4.text_input("Currency", data.get("currency"), disabled=not can_edit() or is_locked)
        
        # --- 📅 INVOICE TIMELINE ---
        with st.expander("📅 Invoice Lifecycle Timeline", expanded=False):
            timeline_events = []
            
            if data.get("created_at"):
                timeline_events.append({
                    "time": data.get("created_at"),
                    "stage": "UPLOADED",
                    "description": f"📤 Uploaded by {data.get('created_by', 'Unknown')}"
                })
            
            if data.get("id"):
                edits = fetch_invoice_edits(data.get("id"))
                if edits:
                    first_edit_time = edits[0].get("edited_at")
                    timeline_events.append({
                        "time": first_edit_time,
                        "stage": "EDITED",
                        "description": f"✏️ {len(edits)} edit(s) made by AP Clerk"
                    })
            
            if data.get("reviewed_by"):
                review_time = data.get("approval_timestamp") or data.get("created_at")
                timeline_events.append({
                    "time": review_time,
                    "stage": "REVIEWED",
                    "description": f"🔍 Reviewed by Manager ({data.get('reviewed_by')})"
                })
            
            if data.get("approval_stage") == "APPROVED":
                timeline_events.append({
                    "time": data.get("approval_timestamp", datetime.now().isoformat()),
                    "stage": "APPROVED",
                    "description": f"✅ Approved by {data.get('approved_by')}"
                })
            elif data.get("approval_stage") == "REJECTED":
                timeline_events.append({
                    "time": data.get("approval_timestamp", datetime.now().isoformat()),
                    "stage": "REJECTED",
                    "description": "🔴 Rejected by Manager"
                })
            
            if data.get("approval_stage") == "AUDITED":
                audit_time = data.get("reprocessed_at") or data.get("approval_timestamp") or datetime.now().isoformat()
                timeline_events.append({
                    "time": audit_time,
                    "stage": "AUDITED",
                    "description": "🟣 Audited and Verified"
                })
            
            if timeline_events:
                for event in timeline_events:
                    event_time = str(event['time'])[:16].replace("T", " ") if event.get('time') else "Unknown"
                    st.write(f"{event_time} — **{event['description']}**")
            else:
                st.info("No lifecycle events yet")

        with st.expander("🤖 AI Field Explanations"):
            explanations = data.get("explanations", {})
            if explanations:
                for field, explanation in explanations.items():
                    readable = field.replace('_', ' ').title()
                    st.write(f"**{readable}**: {explanation}")
            else:
                st.info("No detailed explanations returned.")

        st.write("📦 **Line Items Extraction**")
        items_df = pd.DataFrame(data.get("line_items", []))
        if can_edit() and not is_locked:
            edited_df = st.data_editor(items_df, num_rows="dynamic", use_container_width=True, key="editor")
        else:
            st.dataframe(items_df, use_container_width=True)
            edited_df = items_df 

        st.markdown("### 💰 Total Comparison")
        math_valid, calc_sum, diff = validate_math(extracted_total, edited_df)
        tc1, tc2, tc3 = st.columns(3)
        tc1.metric("AI Extracted Total", f"${extracted_total:,.2f}")
        tc2.metric("Calculated Line Item Total", f"${calc_sum:,.2f}")
        if not math_valid:
            tc3.metric("Difference", f"${diff:,.2f}", delta="Mismatch", delta_color="inverse")
        else:
            tc3.metric("Difference", "$0.00", delta="Matched")

        st.markdown("### 🚦 Risk Analysis")
        duplicate_found = is_duplicate(vendor, date, extracted_total, exclude_id=data.get("id"))
        
        risk_score = 0
        risk_reasons = []
        confidence = data.get("confidence_score", 1.0)
        
        if confidence < 0.7:
            risk_score += 20
            risk_reasons.append(f"Low AI Confidence ({int(confidence*100)}%)")
        if not math_valid:
            risk_score += 30
            risk_reasons.append(f"Math Mismatch (Diff: {diff})")
        if duplicate_found:
            risk_score += 40
            risk_reasons.append("Duplicate Invoice Detected")
        
        vendor_avg = get_vendor_average(vendor)
        if vendor_avg and extracted_total > (vendor_avg * 2):
            risk_score += 25
            risk_reasons.append(f"Amount > 2x Vendor Avg (${vendor_avg:.2f})")
        
        if risk_score >= 60:
            risk_level = "HIGH"
            st.error(f"🚨 **HIGH RISK INVOICE** (Score: {risk_score})")
        elif risk_score >= 30:
            risk_level = "MEDIUM"
            st.warning(f"⚠ **MEDIUM RISK INVOICE** (Score: {risk_score})")
        else:
            risk_level = "LOW"
            st.success(f"✅ **LOW RISK INVOICE** (Score: {risk_score})")

        if risk_reasons:
            st.markdown("**Risk Factors:**")
            for r in risk_reasons:
                st.markdown(f"- {r}")

        status = "Verified"
        flag_reason = None
        if risk_level == "HIGH":
            status = "Flagged"
            flag_reason = f"High Risk: {', '.join(risk_reasons)}"
        elif risk_level == "MEDIUM" and not math_valid:
            status = "Flagged"
            flag_reason = "Math Mismatch"

        st.divider()
        if can_edit() and not is_locked:
            col_left, col_right = st.columns(2)
            if user_role == "AP_CLERK":
                if col_left.button("💾 Save Draft"):
                    save_and_log(vendor, date, extracted_total, currency, edited_df, status, data, flag_reason, user_role, risk_score, risk_level, "UPLOADED")
                if status == "Verified" or status == "Flagged":
                    if col_right.button("📤 Submit for Review"):
                         save_and_log(vendor, date, extracted_total, currency, edited_df, status, data, flag_reason, user_role, risk_score, risk_level, "REVIEWED")
            elif user_role == "FINANCE_MANAGER":
                if col_left.button("💾 Save / Mark Reviewed"):
                     save_and_log(vendor, date, extracted_total, currency, edited_df, status, data, flag_reason, user_role, risk_score, risk_level, "REVIEWED")
                if col_right.button("✅ ACCEPT & SEND TO AUDITOR"):
                    save_and_log(vendor, date, extracted_total, currency, edited_df, "Verified", data, "Manager Accepted - Sent to Auditor", user_role, risk_score, risk_level, "APPROVED")
                if st.button("❌ REJECT INVOICE"):
                     save_and_log(vendor, date, extracted_total, currency, edited_df, "Flagged", data, "Manager Rejected", user_role, risk_score, risk_level, "REJECTED")
        elif user_role == "AUDITOR" and st.session_state.get('audit_mode'):
            st.divider()
            if data.get("approval_stage") == "AUDITED":
                st.success("✅ This invoice has already been audited.")
            else:
                if st.button("✅ Mark as Audited"):
                    save_and_log(vendor, date, extracted_total, currency, edited_df, status, data, "Auditor Verified", user_role, risk_score, risk_level, "AUDITED")
        elif is_locked:
            st.info("🔒 Invoice already approved - editing disabled")
        else:
            st.info("👁️ View Only Mode")

# --- MANAGER REVIEW QUEUE ---
if user_role == "FINANCE_MANAGER" and all_invoices_data:
    st.markdown("---")
    st.markdown("## 🧾 Invoices Awaiting Your Approval")

    review_df = pd.DataFrame(all_invoices_data)
    if 'approval_stage' in review_df.columns:
        review_df = review_df[review_df['approval_stage'] == "REVIEWED"]
        review_df['created_at'] = pd.to_datetime(review_df['created_at'], errors='coerce')
        review_df = review_df.sort_values("created_at")

        if not review_df.empty:
            for i, row in review_df.iterrows():
                with st.container(border=True):
                    col1, col2, col3, col4, col5 = st.columns([2, 1.7, 1.7, 1.3, 1])
                    col1.write(f"**Vendor:**\n{row.get('vendor_name')}")
                    col2.write(f"**Amount:**\n${row.get('total_amount'):,.2f}")
                    col3.write(f"**Date:**\n{row.get('invoice_date')}")
                    col4.write(get_processing_badge(row.get('processing_status')))
                    
                    if col5.button("🔍 Review", key=f"rev_{row['id']}_{i}"):
                        invoice_data = hydrate_invoice_session_data(row)
                        st.session_state['data'] = invoice_data
                        st.session_state['url'] = row.get('file_url')
                        st.session_state['file_bytes'] = None 
                        st.rerun()
        else:
            st.success("No invoices pending your approval 🎉")

# --- MANAGER INVOICE OVERVIEW ---
if user_role == "FINANCE_MANAGER" and all_invoices_data:
    st.markdown("---")
    st.markdown("## 📑 Manager Invoice Overview")
    
    mgr_df = pd.DataFrame(all_invoices_data)
    if 'created_at' in mgr_df.columns:
        mgr_df['created_at'] = pd.to_datetime(mgr_df['created_at'], errors='coerce')
        mgr_df = mgr_df.sort_values('created_at')
    
    tab1, tab2 = st.tabs(["🧾 Received From Clerk", "📁 Approved History"])
    
    with tab1:
        active = mgr_df[mgr_df['approval_stage'] == "REVIEWED"]
        if not active.empty:
            st.subheader(f"Received From Clerk ({len(active)} invoices)")
            for i, row in active.iterrows():
                with st.container(border=True):
                    col1, col2, col3, col4, col5, col6 = st.columns([2, 1.3, 1.3, 1.3, 1.2, 1])
                    col1.write(f"**{row.get('vendor_name')}**")
                    col2.write(f"📅 {row.get('invoice_date')}")
                    col3.write(f"💰 ${row.get('total_amount'):,.2f}")
                    col4.write(get_risk_badge(row.get('risk_level', 'LOW')))
                    col5.write(get_stage_badge(row.get('approval_stage', 'UPLOADED')))
                    if col6.button("🔍 Review", key=f"mgr_pending_{row['id']}_{i}"):
                        invoice_data = hydrate_invoice_session_data(row)
                        st.session_state['data'] = invoice_data
                        st.session_state['url'] = row.get('file_url')
                        st.session_state['file_bytes'] = None
                        st.rerun()
        else:
            st.success("✅ No clerk-submitted invoices pending manager review")
    
    with tab2:
        approved = mgr_df[mgr_df['approval_stage'] == "APPROVED"]
        if not approved.empty:
            st.subheader(f"Approved History ({len(approved)} invoices)")
            for i, row in approved.iterrows():
                with st.container(border=True):
                    col1, col2, col3, col4, col5, col6 = st.columns([2, 1.5, 1.5, 1.2, 1.3, 1])
                    col1.write(f"**{row.get('vendor_name')}**")
                    col2.write(f"📅 {row.get('invoice_date')}")
                    col3.write(f"💰 ${row.get('total_amount'):,.2f}")
                    col4.write(f"✅ {row.get('approved_by', 'Manager')}")
                    col5.write(get_processing_badge(row.get('processing_status')))
                    col6.write(f"🟢 Approved")
        else:
            st.info("No approved invoices yet")

# --- AUDITOR WORK QUEUE ---
if user_role == "AUDITOR" and all_invoices_data:
    st.markdown("---")
    st.markdown("## 📋 Auditor Review Dashboard")
    
    audit_df = pd.DataFrame(all_invoices_data)
    if 'created_at' in audit_df.columns:
        audit_df['created_at'] = pd.to_datetime(audit_df['created_at'], errors='coerce')
        audit_df = audit_df.sort_values('created_at')
    
    tab1, tab2 = st.tabs(["🆕 To Be Audited", "📁 Audited"])
    
    with tab1:
        is_approved = audit_df['approval_stage'] == "APPROVED"
        not_audited = audit_df['audited'] != True if 'audited' in audit_df.columns else True
        pending_audit = audit_df[is_approved & not_audited]
        
        if not pending_audit.empty:
            st.subheader(f"Pending Audit ({len(pending_audit)} invoices)")
            for i, row in pending_audit.iterrows():
                with st.container(border=True):
                    col1, col2, col3, col4, col5 = st.columns([2, 1.5, 1.5, 1.2, 1])
                    col1.write(f"**{row.get('vendor_name')}**")
                    col2.write(f"📅 {row.get('invoice_date')}")
                    col3.write(f"💰 ${row.get('total_amount'):,.2f}")
                    col4.write(get_risk_badge(row.get('risk_level', 'LOW')))
                    if col5.button("👁️ View", key=f"audit_view_{row['id']}_{i}"):
                        invoice_data = hydrate_invoice_session_data(row)
                        st.session_state['data'] = invoice_data
                        st.session_state['url'] = row.get('file_url')
                        st.session_state['file_bytes'] = None
                        st.session_state['audit_mode'] = True
                        st.rerun()
        else:
            st.success("✅ No invoices pending audit - all caught up!")
    
    with tab2:
        audited = audit_df[audit_df['approval_stage'] == "AUDITED"]
        if not audited.empty:
            st.subheader(f"Audited ({len(audited)} invoices)")
            for i, row in audited.iterrows():
                with st.container(border=True):
                    col1, col2, col3, col4, col5 = st.columns([2, 1.5, 1.5, 1.2, 1])
                    col1.write(f"**{row.get('vendor_name')}**")
                    col2.write(f"📅 {row.get('invoice_date')}")
                    col3.write(f"💰 ${row.get('total_amount'):,.2f}")
                    col4.write(f"✅ {row.get('risk_level', 'LOW')}")
                    col5.write("🟣 Audited")
        else:
            st.info("No audited invoices yet")

# --- OPS DASHBOARD ---
if user_role in ["FINANCE_MANAGER", "AUDITOR"]:
    st.markdown("---")
    st.subheader("📈 Operations Control Center")

    if all_invoices_data:
        df = pd.DataFrame(all_invoices_data)
        if 'created_at' in df.columns:
            df['created_at'] = pd.to_datetime(df['created_at']).dt.tz_localize(None)
        
        for col in ['confidence_score', 'risk_score', 'total_amount']:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')

        st.markdown("### 🚨 Operational Alerts")
        alerts = []
        
        if 'risk_level' in df.columns and 'approval_stage' in df.columns:
            high_risk_pending = df[(df['risk_level'] == "HIGH") & (df['approval_stage'].isin(["UPLOADED", "REVIEWED"]))]
            if len(high_risk_pending) > 0:
                alerts.append(f"🔴 **{len(high_risk_pending)} HIGH RISK** invoices require attention")

        if 'confidence_score' in df.columns:
            low_conf = df[df['confidence_score'] < 0.7]
            if len(low_conf) > 0:
                alerts.append(f"⚠ **{len(low_conf)}** invoices have AI confidence below 70%")

        if 'approval_stage' in df.columns:
            now = pd.Timestamp.now()
            df['days_pending'] = (now - df['created_at']).dt.days
            df.loc[df['approval_stage'] == "APPROVED", 'days_pending'] = 0
            
            sla_breach_review = df[(df['approval_stage'] == "REVIEWED") & (df['days_pending'] > 3)]
            if len(sla_breach_review) > 0:
                alerts.append(f"⏱ **{len(sla_breach_review)} SLA BREACHES:** Invoices in review > 72 hours")

        if alerts:
            for alert in alerts:
                st.warning(alert)
        else:
            st.success("✅ No critical operational risks detected.")

        st.markdown("---")
        st.markdown("### ⏱ SLA Performance Metrics")
        
        if 'approval_timestamp' in df.columns:
            df['approval_timestamp'] = pd.to_datetime(df['approval_timestamp'], errors='coerce').dt.tz_localize(None)
            approved_df = df[df['approval_stage'] == "APPROVED"].copy()
            
            if not approved_df.empty:
                approved_df['approval_time_hours'] = (
                    approved_df['approval_timestamp'] - approved_df['created_at']
                ).dt.total_seconds() / 3600
                avg_approval_time = approved_df['approval_time_hours'].mean()
                
                c1, c2, c3 = st.columns(3)
                delta_color = "inverse" if avg_approval_time > 48 else "normal"
                c1.metric("Avg Approval Time", f"{avg_approval_time:.1f} hrs", delta_color=delta_color)
                c2.metric("Fastest Approval", f"{approved_df['approval_time_hours'].min():.1f} hrs")
                c3.metric("Slowest Approval", f"{approved_df['approval_time_hours'].max():.1f} hrs")
            else:
                st.info("No approved invoices yet to calculate SLA metrics.")
        
        st.markdown("### 📊 Analytics Overview")
        chart1, chart2 = st.columns(2)
        with chart1:
            st.markdown("#### 🚨 Risk Distribution")
            if 'risk_level' in df.columns:
                st.bar_chart(df['risk_level'].value_counts(), color="#FF4B4B")
        with chart2:
            st.markdown("#### 🔄 Approval Funnel")
            if 'approval_stage' in df.columns:
                stage_counts = df['approval_stage'].value_counts().reindex(
                    ["UPLOADED", "REVIEWED", "APPROVED", "REJECTED"], fill_value=0
                )
                st.bar_chart(stage_counts)

        chart3, chart4 = st.columns(2)
        with chart3:
            st.markdown("#### 🏢 Top Vendors")
            if 'vendor_name' in df.columns:
                top_vendors = df.groupby("vendor_name")["total_amount"].sum().sort_values(ascending=False).head(5)
                st.bar_chart(top_vendors)
        with chart4:
            st.markdown("#### 🤖 AI Confidence Trend")
            if 'confidence_score' in df.columns:
                st.line_chart(df.set_index("created_at").resample("D")["confidence_score"].mean())

        st.markdown("---")
        st.markdown("### 📄 Export Reports")
        
        # ✅ DOWNLOAD ALL INVOICES (EXCEL)
        st.subheader("📥 Download All Invoices")
        excel_all = export_to_excel(df)
        st.download_button(
            label="📊 Download All Invoices (EXCEL)",
            data=excel_all,
            file_name=f"all_invoices_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        st.divider()
        st.subheader("📋 Filtered Reports (CSV & EXCEL)")
        
        exp1, exp2 = st.columns(2)
        if 'approval_stage' in df.columns:
            approved_report = df[df['approval_stage'] == "APPROVED"]
            csv_approved = approved_report.to_csv(index=False).encode('utf-8')
            exp1.download_button("📥 Download Approved Invoices (CSV)", csv_approved, "approved_invoices.csv", "text/csv")
            
            excel_approved = export_to_excel(approved_report, "approved_invoices")
            exp2.download_button(
                "📊 Download Approved Invoices (EXCEL)", 
                excel_approved, 
                f"approved_invoices_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        exp3, exp4 = st.columns(2)
        risk_cols = ['vendor_name', 'total_amount', 'risk_score', 'risk_level', 'confidence_score', 'flag_reason']
        valid_risk = [c for c in risk_cols if c in df.columns]
        risk_report = df[valid_risk]
        csv_risk = risk_report.to_csv(index=False).encode('utf-8')
        exp3.download_button("📥 Download Risk Report (CSV)", csv_risk, "risk_report.csv", "text/csv")
        
        excel_risk = export_to_excel(risk_report, "risk_report")
        exp4.download_button(
            "📊 Download Risk Report (EXCEL)",
            excel_risk,
            f"risk_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        st.markdown("### 📋 Recent Transactions")
        if 'days_pending' in df.columns:
            df['sla_status'] = df['days_pending'].apply(lambda x: "🚨 BREACH" if x > 3 else "On Track")
        
        display_cols = ["created_at", "vendor_name", "total_amount", "risk_level", "approval_stage", "sla_status"]
        valid_cols = [c for c in display_cols if c in df.columns]
        st.dataframe(df[valid_cols].head(10).style.applymap(lambda x: "color: red; font-weight: bold;" if x == "🚨 BREACH" else "", subset=['sla_status']), use_container_width=True)
    else:
        st.info("No data available.")
else:
    # AP_CLERK: Show simplified dashboard
    st.markdown("---")
    st.subheader("📊 Clerk Work Dashboard")
    if all_invoices_data:
        df = pd.DataFrame(all_invoices_data)

        if 'source' not in df.columns:
            df['source'] = df.apply(get_invoice_source, axis=1)

        manual_df = df[df['source'] == "MANUAL_UPLOAD"]
        ingested_df = df[df['source'] == "INGESTED_EMAIL"]
        clerk_queue = df[df['approval_stage'] == "UPLOADED"] if 'approval_stage' in df.columns else pd.DataFrame()
        submitted_to_manager = df[df['approval_stage'] == "REVIEWED"] if 'approval_stage' in df.columns else pd.DataFrame()
        manager_queue = df[df['approval_stage'] == "REVIEWED"] if 'approval_stage' in df.columns else pd.DataFrame()

        col1, col2, col3, col4 = st.columns(4)
        col1.metric("Manual Uploads", len(manual_df))
        col2.metric("Ingested (IMAP)", len(ingested_df))
        col3.metric("Pending Clerk Verification", len(clerk_queue))
        col4.metric("To Manager", len(manager_queue))

        st.markdown("### 🧾 Clerk Verification Queue")
        if not clerk_queue.empty:
            for i, row in clerk_queue.iterrows():
                with st.container(border=True):
                    c1, c2, c3, c4, c5, c6 = st.columns([2, 1.4, 1.4, 1.3, 1.2, 1])
                    c1.write(f"**{row.get('vendor_name')}**")
                    c2.write(f"📅 {row.get('invoice_date')}")
                    c3.write(f"💰 ${row.get('total_amount'):,.2f}")
                    c4.write(get_processing_badge(row.get('processing_status')))
                    c5.write(get_risk_badge(row.get('risk_level', 'LOW')))
                    if c6.button("🛠️ Open", key=f"clerk_open_{row['id']}_{i}"):
                        invoice_data = hydrate_invoice_session_data(row)
                        st.session_state['data'] = invoice_data
                        st.session_state['url'] = row.get('file_url')
                        st.session_state['file_bytes'] = None
                        st.rerun()
        else:
            st.success("✅ No invoices waiting for clerk verification")

        st.markdown("### 📤 Submitted To Manager")
        if not submitted_to_manager.empty:
            cols = [c for c in ["vendor_name", "invoice_date", "total_amount", "approval_stage", "processing_status", "reviewed_by"] if c in submitted_to_manager.columns]
            st.dataframe(submitted_to_manager[cols].head(30), use_container_width=True)
        else:
            st.info("No invoices submitted to manager yet.")

        tab_manual, tab_ingested = st.tabs(["🖐️ Manual", "📩 IMAP Ingested"])
        with tab_manual:
            if not manual_df.empty:
                cols = [c for c in ["vendor_name", "invoice_date", "total_amount", "approval_stage", "processing_status"] if c in manual_df.columns]
                st.dataframe(manual_df[cols].head(20), use_container_width=True)
            else:
                st.info("No manual uploads yet.")

        with tab_ingested:
            if not ingested_df.empty:
                cols = [c for c in ["vendor_name", "invoice_date", "total_amount", "approval_stage", "processing_status"] if c in ingested_df.columns]
                st.dataframe(ingested_df[cols].head(20), use_container_width=True)
            else:
                st.info("No IMAP-ingested invoices yet.")

        st.caption("Pipeline: Clerk verifies (UPLOADED) → submits to manager (REVIEWED) → manager accepts (APPROVED) → auditor audits (AUDITED).")
    else:
        st.info("No invoices uploaded yet. Use the sidebar to upload your first invoice!")